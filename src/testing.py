from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from openpyxl import Workbook
import os
import platform, urllib.request, openpyxl, operator
import getpass

class Instagram:
    def __init__(self,tag,limit):
        self.tag        = tag
        self.limit      = limit
        self.url        = "https://www.instagram.com"
        self.driver     = webdriver.Chrome(executable_path="/home/local/DEVNET/boskicl/scripts/data_scrapper/src/chromedriver") # path to chromedriver

    def create_dir(self, dirname):
        if os.path.exists("data") and os.path.exists("data/data_" + dirname) and os.path.exists("data/data_" + dirname + '/img'):       # If path exists = pass
            pass
        else:
            if not os.path.exists("data"):
                try:
                    os.mkdir("data")
                    print("Created data directory.")
                except:
                    print("Unable to create data directory: Directory already exists.")
            else:
                print("Unable to create data directory: Directory already exists.")

            if not os.path.exists("data/data_" + dirname):
                try:
                    os.mkdir("data/data_" + dirname)
                    print("Created data/data_{0} directory".format(dirname))
                except:
                    print("Unable to create data/data_{0} directory: Directory already exists.".format(dirname))
            else:
                print("Unable to create data/data_{0} directory: Directory already exists.".format(dirname))

            if not os.path.exists("data/data_" + dirname + '/img'):
                try:
                    os.mkdir("data/data_" + dirname + '/img')
                    print("Created data/data_{0}/img directory.".format(dirname))
                except:
                    print("Unable to create data/data_{0}/img directory: Directory already exists.".format(dirname))
            else:
                print("Unable to create data/data_{0}/img directory: Directory already exists.".format(dirname))

    # def excel_exporter(self,names, comments, file_path):
    #     fp = file_path
    #     tag_File = fp + "/" + self.tag + "_Instagram.xlsx"
    #     temp = {}
    #     temp_names = []
    #     temp_comments = []
    #     if os.path.isfile(fname):
    #         saved = pd.read_excel(fname)
    #         temp_names.extend(saved['name'])
    #         temp_comments.extend(saved['comment'])
    #     temp_names.extend(names)
    #     temp_comments.extend(comments)
    #     temp.update({'name': temp_names, 'comment': temp_comments})
    #     df = pd.DataFrame(temp)
    #     writer = ExcelWriter(fname)
    #     df.to_excel(writer, 'ridwan kamil', index=False)
    #     writer.save()

    def Tag_Scrapper(self):
        #directory   = self.create_dir(self.tag)                                 # Create directory with hashtag
        driver = self.driver
        driver.get('{0}/explore/tags/{1}'.format(self.url,self.tag))
        print("Loading Posts")
        sleep(15)
        print("Loading Data")

        file_path = "data/data_" + self.tag
        keyword = self.tag

        actions = ActionChains(driver)
        actions.send_keys(Keys.SPACE).perform()
        actions.send_keys(Keys.SPACE).perform()
        actions.send_keys(Keys.SPACE).perform()
        sleep(5)

        clear = lambda: os.system('cls')
        msg = "Loading Images"
        class_div_img = ["_si7dy"]
        for div in class_div_img:
            if len(driver.find_elements_by_class_name(div)) > 1:
                while (len(driver.find_elements_by_class_name(div)) ) <= self.limit :
                    actions.send_keys(Keys.SPACE).perform()
                    msg = msg + "."
                    print(msg)
                    print(len(driver.find_elements_by_class_name(div)))
                    sleep(2.5)
                    if len(msg) > 18:
                        msg = "Loading Images"
        print(str(self.limit) + " images loaded")

        driver.find_element_by_class_name('eLAPa').click()

        # img_src = []
        # img_caption = []
        # hashtags = {}
        #
        # for data in driver.find_elements_by_class_name("FFVAD"):
        #     #     u.get_attribute("href").split("/")[4]
        #     img_caption.append(data.get_attribute("alt"))
        #     img_src.append(data.get_attribute("src"))
        #
        #
        # #print('Skipping first 9 photos: Going to most recent.')
        # img_caption = img_caption[9:self.limit + 9]
        # img_src = img_src[9:self.limit + 9]
        # img_caption.sort()
        #
        # tag_File = file_path + "/" + self.tag + "_Instagram.xlsx"
        # wb = Workbook()
        # ws_Captions = wb.create_sheet(title="Caption")
        # col = 'A'
        # row = 1
        # print('Line 105')
        # print("Dumping data in excel file")
        # for caption in img_caption:
        #     tags = caption.split("#")
        #     # write caption to excel file
        #     ws_Captions[col + str(row)] = tags[0]
        #     row += 1
        #
        #     # strip tags
        #     tags = tags[1:]
        #     for self.tag in tags:
        #         cleaned = self.tag.replace(" ", "").replace("\n", "")
        #         cleaned = cleaned.lower()
        #         print(cleaned)
        #         if cleaned not in hashtags and len(cleaned) < 20:
        #             hashtags[cleaned] = 1
        #         elif cleaned in hashtags and len(cleaned) < 20:
        #             hashtags[cleaned] = hashtags[cleaned] + 1
        #
        # # sort hashtags with frequencies and store them in excel
        # hashtags = sorted(hashtags.items(), key=operator.itemgetter(1), reverse=True)
        #
        # ws_Tags = wb.create_sheet(title="Tags")
        # tagName = 'A'
        # tagFreq = 'B'
        # row = 1
        #
        # for self.tag in hashtags:
        #     ws_Tags[tagName + str(row)] = self.tag[0]
        #     ws_Tags[tagFreq + str(row)] = self.tag[1]
        #     row += 1
        #
        # wb.save(tag_File)
        #
        # print("Dumping Images. This will take some time!")
        # row = 1
        # for src in img_src:
        #     urllib.request.urlretrieve(src, file_path + '/img/Instagram_' + str(row) + ".jpeg")
        #     row += 1
        #     if (row % 10 == 0):
        #         print("(" + str(row) + "/" + str(len(img_src)) + ") Images Downloaded")
        #
        # print("Closing Instagram")
        # driver.quit()
insta = Instagram('tree',1)
insta.Tag_Scrapper()
